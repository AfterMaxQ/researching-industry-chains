"""Runner JSON、review 与九字段 XLSX 的端到端验收。"""

from __future__ import annotations

import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from industry_chain_skills.errors import ClientError
from industry_chain_skills.excel import HEADERS
from industry_chain_skills.review import ReviewService
from industry_chain_skills.runner import RunnerService
from industry_chain_skills.source_service import SourceService
from industry_chain_skills.storage import RunnerStore
from industry_chain_skills.work import WorkService


NOW = datetime(2026, 9, 3, tzinfo=timezone.utc)


def accept_result(label: str, url: str) -> dict:
    """创建包含父节点、子节点和两个企业的正式来源。"""
    return {
        "outcome": "accept",
        "source": {"name": f"{label}研究院", "url": url},
        "description": f"{label}最终来源说明",
        "chain": [
            {
                "name": "上游",
                "children": [
                    {
                        "name": f"{label}材料",
                        "companies": [f"{label}甲公司", f"{label}乙公司"],
                    }
                ],
            }
        ],
    }


def review_result(label: str, url: str, empty_chain: bool = False) -> dict:
    """创建来源级待审结果。"""
    return {
        "outcome": "review",
        "source": {"name": f"{label}研究院", "url": url},
        "description": f"{label}仍需人工确认",
        "chain": [] if empty_chain else [{"name": "上游"}],
        "uncertainties": [
            {
                "message": f"{label}结构未可靠闭环",
                "evidence": [
                    {
                        "locator": "PDF 第 17 页图 5",
                        "description": "图中连接关系需要复核。",
                    }
                ],
            }
        ],
    }


class HitlIntegrationTests(unittest.TestCase):
    """验证正式数据和审核数据在文件事务中保持分离。"""

    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.runs_root = Path(self.tempdir.name)
        self.store = RunnerStore(self.runs_root)
        runner = RunnerService(self.store, clock=lambda: NOW)
        created = runner.create("HITL 端到端", topic="锡膏")
        self.runner_id = created["runner_id"]
        tokens = iter([f"token-{index}" for index in range(1, 20)])
        review_ids = iter([f"review_{index}" for index in range(1, 20)])
        self.work = WorkService(
            self.store,
            clock=lambda: NOW,
            token_factory=lambda: next(tokens),
        )
        self.source = SourceService(
            self.store,
            clock=lambda: NOW,
            review_id_factory=lambda: next(review_ids),
        )
        self.review = ReviewService(self.store, clock=lambda: NOW)

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    @property
    def workbook_path(self) -> Path:
        """返回当前 Runner 的交付工作簿。"""
        return (
            self.runs_root
            / self.runner_id
            / f"{self.runner_id}_交付数据.xlsx"
        )

    def workbook_values_and_links(self) -> tuple[list[tuple], list[str | None]]:
        """读取 XLSX 全部单元格值和 URL 超链接。"""
        workbook = load_workbook(self.workbook_path)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        links = []
        for index in range(2, sheet.max_row + 1):
            hyperlink = sheet.cell(row=index, column=8).hyperlink
            links.append(hyperlink.target if hyperlink else None)
        workbook.close()
        return values, links

    def test_multi_source_review_round_trip_and_xlsx_projection(self) -> None:
        topic_work = self.work.claim_next(self.runner_id)
        self.source.submit(
            self.runner_id,
            topic_work["work_id"],
            topic_work["claim_token"],
            accept_result("A", "https://example.com/a"),
        )
        queued = self.source.submit(
            self.runner_id,
            topic_work["work_id"],
            topic_work["claim_token"],
            review_result("B", "https://example.com/b"),
        )
        self.source.submit(
            self.runner_id,
            topic_work["work_id"],
            topic_work["claim_token"],
            accept_result("C", "https://example.com/c"),
        )

        done = self.work.done(
            self.runner_id,
            topic_work["work_id"],
            topic_work["claim_token"],
        )
        self.assertEqual("awaiting_review", done["topic"]["status"])
        before_values, before_links = self.workbook_values_and_links()
        self.assertEqual(tuple(HEADERS), before_values[0])
        before_sources = [row[1] for row in before_values[1:]]
        self.assertEqual(["A研究院", "A研究院", "C研究院", "C研究院"], before_sources)
        self.assertNotIn("B研究院", before_sources)
        self.assertEqual(
            ["https://example.com/a"] * 2 + ["https://example.com/c"] * 2,
            before_links,
        )

        self.review.return_to_agent(
            self.runner_id,
            queued["review_item_id"],
            queued["version"],
        )
        review_work = self.work.claim_next(self.runner_id)
        accepted = self.source.submit(
            self.runner_id,
            review_work["work_id"],
            review_work["claim_token"],
            accept_result("B", "https://example.com/b"),
        )

        self.assertEqual(queued["review_item_id"], accepted["review_item_id"])
        state = self.store.read(self.runner_id)
        topic = state["topics"][0]
        self.assertEqual("completed", topic["status"])
        self.assertEqual(1, len(topic["review_items"]))
        final_values, final_links = self.workbook_values_and_links()
        self.assertEqual(7, len(final_values))
        self.assertEqual("B最终来源说明", final_values[5][8])
        self.assertEqual("B甲公司、B乙公司", final_values[6][6])
        self.assertEqual("https://example.com/b", final_links[-1])
        self.assertEqual(
            {"runner.json", f"{self.runner_id}_交付数据.xlsx"},
            {path.name for path in (self.runs_root / self.runner_id).iterdir()},
        )

    def test_empty_chain_review_rejects_to_no_qualified_source(self) -> None:
        topic_work = self.work.claim_next(self.runner_id)
        queued = self.source.submit(
            self.runner_id,
            topic_work["work_id"],
            topic_work["claim_token"],
            review_result("空树", "https://example.com/empty", empty_chain=True),
        )
        self.work.done(
            self.runner_id,
            topic_work["work_id"],
            topic_work["claim_token"],
        )

        self.review.reject(
            self.runner_id,
            queued["review_item_id"],
            queued["version"],
        )

        state = self.store.read(self.runner_id)
        self.assertEqual("no_qualified_source", state["topics"][0]["status"])
        values, links = self.workbook_values_and_links()
        self.assertEqual([tuple(HEADERS)], values)
        self.assertEqual([], links)

    def test_duplicate_approval_failure_keeps_review_json_and_xlsx_unchanged(self) -> None:
        topic_work = self.work.claim_next(self.runner_id)
        self.source.submit(
            self.runner_id,
            topic_work["work_id"],
            topic_work["claim_token"],
            accept_result("正式", "https://example.com/duplicate"),
        )
        queued = self.source.submit(
            self.runner_id,
            topic_work["work_id"],
            topic_work["claim_token"],
            review_result("待审", "https://example.com/duplicate"),
        )
        self.work.done(
            self.runner_id,
            topic_work["work_id"],
            topic_work["claim_token"],
        )
        before_state = self.store.read(self.runner_id)
        before_values, before_links = self.workbook_values_and_links()

        with self.assertRaises(ClientError) as error:
            self.review.approve(
                self.runner_id,
                queued["review_item_id"],
                queued["version"],
                "人工最终说明",
                [{"name": "上游", "companies": ["待审甲公司"]}],
            )

        self.assertEqual("SOURCE_GROUP_DUPLICATE_URL", error.exception.code)
        self.assertEqual(before_state, self.store.read(self.runner_id))
        self.assertEqual(
            (before_values, before_links),
            self.workbook_values_and_links(),
        )


if __name__ == "__main__":
    unittest.main()
