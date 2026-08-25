"""九字段交付工作簿投影。"""

from pathlib import Path
from typing import Iterator

from openpyxl import Workbook


HEADERS = (
    "主题",
    "信源主体",
    "分类1",
    "分类2",
    "分类3",
    "分类4",
    "公司",
    "信源URL",
    "备注",
)


def iter_records_in_global_order(state: dict) -> Iterator[dict]:
    """按来源组全局顺序和组内行顺序遍历记录。"""
    groups = [
        group
        for topic in state["topics"]
        for group in topic.get("source_groups", [])
    ]
    for group in sorted(groups, key=lambda item: item["order"]):
        for row in sorted(group["rows"], key=lambda item: item["order"]):
            yield row["record"]


def write_workbook(state: dict, target: Path) -> None:
    """从 Runner 当前状态生成九列交付工作簿。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "交付数据"
    sheet.append(HEADERS)

    for record in iter_records_in_global_order(state):
        sheet.append([record[header] for header in HEADERS])
        url_cell = sheet.cell(sheet.max_row, 8)
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.style = "Hyperlink"

    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
