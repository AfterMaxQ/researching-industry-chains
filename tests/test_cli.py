"""产业链 CLI 端到端测试。"""

import json
import subprocess
import sys

from openpyxl import load_workbook


def run_cli(*args: str, input_text: str | None = None) -> dict:
    """运行 CLI 并校验统一 JSON 外壳。"""
    result = subprocess.run(
        [sys.executable, "-m", "industry_chain_skills.cli", *args],
        input=input_text,
        text=True,
        capture_output=True,
        check=False,
    )
    assert result.stderr == ""
    payload = json.loads(result.stdout)
    assert payload["ok"] is (result.returncode == 0)
    return payload


def test_cli_create_claim_insert_finish_and_export(tmp_path, topic_config) -> None:
    """CLI 完成创建、领取、写入、结束和导出闭环。"""
    runs_root = tmp_path / "runs"
    source_json = tmp_path / "source.json"
    source_json.write_text(
        json.dumps({
            "records": [{
                "主题": "测试主题",
                "信源主体": "测试研究院",
                "分类1": "上游",
                "分类2": "材料",
                "分类3": "",
                "分类4": "",
                "公司": "甲公司",
                "信源URL": "https://example.com/report",
                "备注": "发布日期未识别",
            }]
        }, ensure_ascii=False),
        encoding="utf-8",
    )
    created = run_cli(
        "--runs-root", str(runs_root),
        "runner", "create", "--name", "批次", "--config", str(topic_config),
    )["data"]
    claimed = run_cli(
        "--runs-root", str(runs_root),
        "topic", "claim-next", "--runner-id", created["runner_id"],
    )["data"]
    inserted = run_cli(
        "--runs-root", str(runs_root),
        "dataset", "insert", "--runner-id", created["runner_id"],
        "--scope", "source_group", "--parent-id", claimed["node_id"],
        "--claim-token", claimed["claim_token"], "--input", str(source_json),
    )["data"]
    assert inserted["source_group_id"].startswith("source_")

    finished = run_cli(
        "--runs-root", str(runs_root),
        "topic", "finish", "--runner-id", created["runner_id"],
        "--node-id", claimed["node_id"], "--claim-token", claimed["claim_token"],
        "--outcome", "completed",
    )["data"]
    assert finished["status"] == "completed"
    status = run_cli(
        "--runs-root", str(runs_root),
        "runner", "status", "--runner-id", created["runner_id"],
    )["data"]
    assert status["counts"]["completed"] == 1
    workbook = load_workbook(status["xlsx_path"])
    assert workbook["交付数据"].cell(2, 8).hyperlink.target == "https://example.com/report"


def test_concurrent_claim_next_has_one_winner(tmp_path) -> None:
    """同一待处理主题只有一个并发领取者成功。"""
    runs_root = tmp_path / "runs"
    topic_config = tmp_path / "one-topic.yaml"
    topic_config.write_text(
        "themes:\n  唯一主题:\n    path: [测试分类, 唯一主题]\n    aliases: []\n",
        encoding="utf-8",
    )
    created = run_cli(
        "--runs-root", str(runs_root),
        "runner", "create", "--name", "并发批次", "--config", str(topic_config),
    )["data"]
    command = [
        sys.executable, "-m", "industry_chain_skills.cli",
        "--runs-root", str(runs_root),
        "topic", "claim-next", "--runner-id", created["runner_id"],
    ]
    processes = [
        subprocess.Popen(command, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        for _ in range(2)
    ]
    responses = []
    for process in processes:
        stdout, stderr = process.communicate(timeout=10)
        assert stderr == ""
        responses.append(json.loads(stdout))

    assert sum(response["ok"] for response in responses) == 1
    failed = next(response for response in responses if not response["ok"])
    assert failed["error"]["code"] == "NO_PENDING_TOPIC"
    winner = next(response["data"] for response in responses if response["ok"])
    topic = run_cli(
        "--runs-root", str(runs_root),
        "topic", "get", "--runner-id", created["runner_id"],
        "--node-id", winner["node_id"],
    )["data"]
    assert topic["claim"]["token"] == winner["claim_token"]
