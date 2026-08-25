"""Runner 存储与工作簿投影测试。"""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from industry_chain_skills.errors import ClientError
from industry_chain_skills.storage import RunnerStore


def make_record(topic: str, url: str) -> dict[str, str]:
    """创建一条完整九字段记录。"""
    return {
        "主题": topic,
        "信源主体": "测试研究院",
        "分类1": "上游",
        "分类2": "材料",
        "分类3": "",
        "分类4": "",
        "公司": "甲公司",
        "信源URL": url,
        "备注": "",
    }


def build_runner_state_with_out_of_order_groups() -> dict:
    """创建来源组顺序与主题顺序不同的 Runner 状态。"""
    timestamp = "2026-08-25T08:00:00+00:00"
    second_group = {
        "source_group_id": "source_second",
        "order": 2,
        "created_at": timestamp,
        "updated_at": timestamp,
        "rows": [{
            "row_id": "row_second",
            "order": 1,
            "created_at": timestamp,
            "updated_at": timestamp,
            "record": make_record("后写入主题", "https://example.com/second"),
        }],
    }
    first_group = {
        "source_group_id": "source_first",
        "order": 1,
        "created_at": timestamp,
        "updated_at": timestamp,
        "rows": [{
            "row_id": "row_first",
            "order": 1,
            "created_at": timestamp,
            "updated_at": timestamp,
            "record": make_record("先写入主题", "https://example.com/first"),
        }],
    }
    return {
        "runner_id": "20260825-080000-test-a1b2c3",
        "name": "test",
        "topic_identity_path": "C:/tmp/topics.yaml",
        "created_at": timestamp,
        "updated_at": timestamp,
        "topics": [
            {"node_id": "node_0001", "主题": "后写入主题", "path": [],
             "aliases": [], "order": 1, "status": "completed",
             "last_error": None, "claim": None, "source_groups": [second_group]},
            {"node_id": "node_0002", "主题": "先写入主题", "path": [],
             "aliases": [], "order": 2, "status": "completed",
             "last_error": None, "claim": None, "source_groups": [first_group]},
        ],
    }


def test_store_writes_json_and_ordered_hyperlink_workbook(tmp_path: Path) -> None:
    """存储层按全局顺序生成带超链接的九列工作簿。"""
    state = build_runner_state_with_out_of_order_groups()
    store = RunnerStore(tmp_path / "runs")
    workbook_path = store.create(state)

    loaded = store.read(state["runner_id"])
    assert loaded["runner_id"] == state["runner_id"]

    workbook = load_workbook(workbook_path)
    sheet = workbook["交付数据"]
    assert [sheet.cell(1, column).value for column in range(1, 10)] == [
        "主题", "信源主体", "分类1", "分类2", "分类3",
        "分类4", "公司", "信源URL", "备注",
    ]
    assert sheet.cell(2, 1).value == "先写入主题"
    assert sheet.cell(2, 8).hyperlink.target == "https://example.com/first"


def test_projection_failure_preserves_existing_pair(tmp_path, monkeypatch) -> None:
    """工作簿投影失败时保留原有 JSON 和 XLSX。"""
    state = build_runner_state_with_out_of_order_groups()
    store = RunnerStore(tmp_path / "runs")
    workbook_path = store.create(state)
    json_path = workbook_path.parent / "runner.json"
    before_json = json_path.read_bytes()
    before_xlsx = workbook_path.read_bytes()

    def fail_projection(state: dict, target: Path) -> None:
        """模拟工作簿写入失败。"""
        raise OSError("投影失败")

    monkeypatch.setattr("industry_chain_skills.storage.write_workbook", fail_projection)
    with pytest.raises(ClientError):
        store.mutate_dataset(
            state["runner_id"], lambda current: current.update(name="changed")
        )

    assert json_path.read_bytes() == before_json
    assert workbook_path.read_bytes() == before_xlsx


def test_state_mutation_does_not_rebuild_workbook(tmp_path, monkeypatch) -> None:
    """仅修改运行状态时不触碰交付工作簿。"""
    state = build_runner_state_with_out_of_order_groups()
    store = RunnerStore(tmp_path / "runs")
    workbook_path = store.create(state)
    before_xlsx = workbook_path.read_bytes()

    def fail_projection(state: dict, target: Path) -> None:
        """若被调用则测试失败。"""
        raise AssertionError("不应重建工作簿")

    monkeypatch.setattr("industry_chain_skills.storage.write_workbook", fail_projection)
    result = store.mutate_state(
        state["runner_id"], lambda current: current.update(name="new") or "ok"
    )

    assert result == "ok"
    assert store.read(state["runner_id"])["name"] == "new"
    assert workbook_path.read_bytes() == before_xlsx
